import argparse
import duckdb
import pandas as pd
import os
import re
import warnings
import hashlib
import sys
import asyncio
from datetime import datetime
from typing import Iterator, List, Any, Optional

# Add project root to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database import DatabaseRepository
import structlog

logger = structlog.get_logger(__name__)

# Suppress openpyxl warnings about data validation
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Streaming import constants
MAX_EXCEL_SIZE = 100 * 1024 * 1024  # 100MB
STREAMING_CHUNK_SIZE = 10_000

def calculate_md5(file_path):
    """Calculate MD5 hash of a file."""
    hash_md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()

def init_meta_table(conn):
    """Initialize metadata table for tracking file imports and usage.

    Creates _data_skill_meta table if it doesn't exist, and adds URL source
    columns if they don't exist (for backward compatibility).
    """
    # Create table if it doesn't exist
    conn.execute('''
        CREATE TABLE IF NOT EXISTS _data_skill_meta (
            file_name TEXT,
            table_name TEXT PRIMARY KEY,
            md5_hash TEXT,
            import_time DATETIME,
            last_used_time DATETIME,
            source_url TEXT,
            source_format TEXT,
            auth_type TEXT,
            last_refresh_time DATETIME
        )
    ''')

    # Check if URL columns exist (for backward compatibility)
    # DuckDB uses information_schema instead of PRAGMA table_info
    rows = conn.execute('''
        SELECT column_name FROM information_schema.columns
        WHERE table_name = '_data_skill_meta'
    ''').fetchall()
    existing_columns = {row[0] for row in rows}

    # Add URL columns if missing
    url_columns = [
        ('source_url', 'TEXT'),
        ('source_format', 'TEXT'),
        ('auth_type', 'TEXT'),
        ('last_refresh_time', 'DATETIME'),
    ]

    for col_name, col_type in url_columns:
        if col_name not in existing_columns:
            try:
                conn.execute(f'ALTER TABLE _data_skill_meta ADD COLUMN {col_name} {col_type}')
            except (duckdb.Error, Exception):
                # Column might already exist from concurrent operation
                pass

    metadata_columns = [
        ('file_path', 'TEXT'),
        ('row_count', 'INTEGER'),
        ('parent_tables', 'TEXT'),
    ]

    for col_name, col_type in metadata_columns:
        if col_name not in existing_columns:
            try:
                conn.execute(f'ALTER TABLE _data_skill_meta ADD COLUMN {col_name} {col_type}')
            except (duckdb.Error, Exception):
                pass

    conn.commit()

def check_duplicate_import(conn, md5_hash):
    """Check if the file has already been imported with the same MD5. Returns a list of table names."""
    results = conn.execute('''
        SELECT table_name FROM _data_skill_meta 
        WHERE md5_hash = ?
    ''', (md5_hash,)).fetchall()
    if results:
        tables = [r[0] for r in results]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for t in tables:
            conn.execute('''
                UPDATE _data_skill_meta 
                SET last_used_time = ? 
                WHERE table_name = ?
            ''', (now, t))
        conn.commit()
        return tables
    return None

def record_import(conn, file_name, table_name, md5_hash, file_path=None, row_count=None):
    """Record import metadata."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute('''
        INSERT OR REPLACE INTO _data_skill_meta
        (file_name, table_name, md5_hash, import_time, last_used_time, file_path, row_count)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (file_name, table_name, md5_hash, now, now, file_path, row_count))
    conn.commit()


def record_url_import(conn, url: str, table_name: str, source_format: str, auth_type: Optional[str] = None):
    """Record URL data source import metadata."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute('''
        INSERT OR REPLACE INTO _data_skill_meta
        (file_name, table_name, import_time, last_used_time,
         source_url, source_format, auth_type, last_refresh_time)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (url, table_name, now, now, url, source_format, auth_type, now))
    conn.commit()


def get_url_sources(conn) -> List[dict]:
    """Get all URL data sources from metadata."""
    rows = conn.execute('''
        SELECT table_name, source_url, source_format, auth_type
        FROM _data_skill_meta
        WHERE source_url IS NOT NULL
    ''').fetchall()
    columns = ['table_name', 'source_url', 'source_format', 'auth_type']
    return [dict(zip(columns, row)) for row in rows]

def clean_column_names(columns):
    """Clean column names to be valid DuckDB identifiers."""
    cleaned = []
    seen = set()
    for col in columns:
        if pd.isna(col):
            col = "Unnamed"
        
        # Convert to string and clean
        col_str = str(col).strip()
        # Replace non-alphanumeric with underscore
        col_str = re.sub(r'\W+', '_', col_str)
        # Remove leading/trailing underscores
        col_str = col_str.strip('_')
        
        if not col_str or col_str.isdigit():
            col_str = f"col_{col_str}"
            
        # Ensure unique
        final_col = col_str
        counter = 1
        while final_col.lower() in [s.lower() for s in seen]:
            final_col = f"{col_str}_{counter}"
            counter += 1
            
        seen.add(final_col)
        cleaned.append(final_col)
    return cleaned

def find_header_row(df, max_rows=10):
    """Find the most likely header row by looking for the row with the most non-null values."""
    best_row_idx = 0
    max_non_nulls = 0
    
    for i in range(min(max_rows, len(df))):
        non_null_count = df.iloc[i].count()
        if non_null_count > max_non_nulls:
            max_non_nulls = non_null_count
            best_row_idx = i
            
    return best_row_idx

def unmerge_and_fill_excel(file_path, sheet_name=None):
    """
    Reads an Excel file, unmerges any merged cells, and fills them with the top-left value.
    Returns a pandas DataFrame.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=False)
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    # Extract merged cells ranges
    merged_ranges = list(sheet.merged_cells.ranges)

    # Iterate over merged cells and fill them with the top-left value
    for merged_cell in merged_ranges:
        min_col, min_row, max_col, max_row = merged_cell.bounds
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value

        # Unmerge the cells (optional, but good for clean structure)
        sheet.unmerge_cells(str(merged_cell))

        # Fill the unmerged cells with the top-left value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col, value=top_left_cell_value)

    # Read into pandas
    data = sheet.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)
    return df


def import_excel_streaming(
    file_path: str,
    db_path: str,
    table_name: str,
    conn,
    drop_null_columns: bool = True
) -> Iterator[int]:
    """Import Excel using streaming (read_only mode).

    Yields row count after each chunk for progress tracking.
    """
    import openpyxl

    # Validate file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_EXCEL_SIZE:
        raise ValueError(f"Excel 文件过大，最大支持 100MB: {file_path}")

    # Open workbook in read-only mode for streaming
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active

    header: Optional[List[str]] = None
    row_buffer: List[tuple] = []
    rows_processed = 0
    first_chunk = True
    null_column_indices: set = set()  # Track columns with only NULL values

    try:
        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows at the start
            if header is None:
                if all(cell is None for cell in row):
                    continue
                # First non-empty row is the header
                header = clean_column_names(list(row))
                null_column_indices = set(range(len(header)))  # All columns potentially null
                continue

            # Skip completely empty data rows
            if all(cell is None for cell in row):
                continue

            # Track non-null values for column analysis
            for i, cell in enumerate(row):
                if cell is not None:
                    null_column_indices.discard(i)

            # Add row to buffer
            row_buffer.append(tuple(row))
            rows_processed += 1

            # Insert when buffer reaches chunk size
            if len(row_buffer) >= STREAMING_CHUNK_SIZE:
                _insert_chunk(conn, table_name, header, row_buffer, first_chunk)
                first_chunk = False
                row_buffer = []
                yield rows_processed

        # Insert remaining rows
        if row_buffer:
            _insert_chunk(conn, table_name, header, row_buffer, first_chunk)
            yield rows_processed

        # Drop all-null columns if requested
        if drop_null_columns and null_column_indices and header:
            _drop_null_columns(conn, table_name, header, null_column_indices)

        conn.commit()

    finally:
        wb.close()


def _drop_null_columns(
    conn,
    table_name: str,
    columns: List[str],
    null_indices: set
) -> None:
    """Remove columns that contain only NULL values."""
    if not null_indices:
        return

    keep_columns = [col for i, col in enumerate(columns) if i not in null_indices]

    if len(keep_columns) == len(columns):
        return

    if not keep_columns:
        return

    temp_table = f"{table_name}_clean"
    col_defs = ", ".join([f'"{col}" TEXT' for col in keep_columns])
    conn.execute(f"CREATE TABLE {temp_table} ({col_defs})")

    cols_str = ", ".join([f'"{col}"' for col in keep_columns])
    conn.execute(f"INSERT INTO {temp_table} SELECT {cols_str} FROM {table_name}")

    conn.execute(f"DROP TABLE {table_name}")
    conn.execute(f"ALTER TABLE {temp_table} RENAME TO {table_name}")


def _insert_chunk(
    conn,
    table_name: str,
    columns: List[str],
    rows: List[tuple],
    is_first_chunk: bool
) -> None:
    """Insert a chunk of rows into the database."""
    if is_first_chunk:
        conn.execute(f"DROP TABLE IF EXISTS {table_name}")
        col_defs = ", ".join([f'"{col}" TEXT' for col in columns])
        conn.execute(f"CREATE TABLE {table_name} ({col_defs})")

    placeholders = ", ".join(["?" for _ in columns])
    insert_sql = f"INSERT INTO {table_name} VALUES ({placeholders})"
    conn.executemany(insert_sql, rows)

def import_to_sqlite(file_path, db_path, table_name=None):
    """Import CSV/XLSX into DuckDB, handling complex headers and merged cells."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    base_name = os.path.splitext(os.path.basename(file_path))[0]

    if not table_name:
        table_name = re.sub(r'\W+', '_', base_name).strip('_')

    repo = DatabaseRepository(db_path)

    with repo.connection() as conn:
        init_meta_table(conn)

        file_md5 = calculate_md5(file_path)
        file_name = os.path.basename(file_path)

        existing_tables = check_duplicate_import(conn, file_md5)
        if existing_tables:
            print(f"File '{file_name}' with identical content already imported as tables: {', '.join(existing_tables)}. Skipping import.")
            return existing_tables

        def get_unique_table_name(base_name):
            t_name = base_name
            counter = 1
            while True:
                result = conn.execute(
                    "SELECT table_name FROM information_schema.tables WHERE table_name = ? AND table_schema = 'main'",
                    (t_name,)
                ).fetchone()
                if not result:
                    return t_name
                t_name = f"{base_name}_v{counter}"
                counter += 1

        def _df_to_duckdb(conn, df, table_name, if_exists='replace'):
            """Write a DataFrame to DuckDB using register + CREATE TABLE AS."""
            conn.register('_tmp_df', df)
            if if_exists == 'replace':
                conn.execute(f"CREATE OR REPLACE TABLE {table_name} AS SELECT * FROM _tmp_df")
            elif if_exists == 'append':
                conn.execute(f"INSERT INTO {table_name} SELECT * FROM _tmp_df")
            conn.unregister('_tmp_df')

        imported_tables = []

        if ext == '.csv':
            target_table = get_unique_table_name(table_name)
            print(f"Importing {file_path} to table '{target_table}' in {db_path}...")

            chunk_size = 50000
            first_chunk = True
            sample_df = pd.read_csv(file_path, nrows=20, header=None)
            header_idx = find_header_row(sample_df)

            for chunk in pd.read_csv(file_path, skiprows=header_idx, chunksize=chunk_size):
                if first_chunk:
                    chunk.columns = clean_column_names(chunk.columns)
                    final_cols = chunk.columns
                    _df_to_duckdb(conn, chunk, target_table, 'replace')
                    first_chunk = False
                else:
                    chunk.columns = final_cols
                    _df_to_duckdb(conn, chunk, target_table, 'append')

            print(f"CSV import completed successfully.")
            record_import(conn, file_name, target_table, file_md5)
            imported_tables.append(target_table)

        elif ext in ['.xlsx', '.xls', '.et']:
            file_size = os.path.getsize(file_path)
            if file_size > MAX_EXCEL_SIZE:
                raise ValueError(f"Excel 文件过大，最大支持 100MB: {file_path}")

            if ext == '.et':
                print(f"Using pandas import for WPS .et file ({file_size / 1024 / 1024:.2f} MB)")
                try:
                    xl = pd.ExcelFile(file_path)
                    sheet_names = xl.sheet_names
                except Exception:
                    sheet_names = [0]

                for sheet_name in sheet_names:
                    if len(sheet_names) > 1:
                        safe_sheet_name = str(sheet_name).strip()
                        safe_sheet_name = re.sub(r'\W+', '_', safe_sheet_name).strip('_')
                        base_sheet_table = f"{table_name}_{safe_sheet_name}"
                    else:
                        base_sheet_table = table_name

                    target_table = get_unique_table_name(base_sheet_table)
                    print(f"Processing Excel sheet '{sheet_name}' to table '{target_table}'...")

                    sample_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=20, header=None)
                    header_idx = find_header_row(sample_df)
                    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=header_idx)

                    df.dropna(how='all', inplace=True)
                    df.dropna(axis=1, how='all', inplace=True)
                    df.columns = clean_column_names(df.columns)
                    _df_to_duckdb(conn, df, target_table, 'replace')
                    print(f"Sheet '{sheet_name}' import completed. Loaded {len(df)} rows.")

                    sheet_file_name = f"{file_name}::{sheet_name}" if len(sheet_names) > 1 else file_name
                    record_import(conn, sheet_file_name, target_table, file_md5)
                    imported_tables.append(target_table)
            else:
                print(f"Using streaming import for Excel file ({file_size / 1024 / 1024:.2f} MB)")

                try:
                    xl = pd.ExcelFile(file_path)
                    sheet_names = xl.sheet_names
                except Exception:
                    sheet_names = [0]

                for sheet_name in sheet_names:
                    if len(sheet_names) > 1:
                        safe_sheet_name = str(sheet_name).strip()
                        safe_sheet_name = re.sub(r'\W+', '_', safe_sheet_name).strip('_')
                        base_sheet_table = f"{table_name}_{safe_sheet_name}"
                    else:
                        base_sheet_table = table_name

                    target_table = get_unique_table_name(base_sheet_table)
                    print(f"Processing Excel sheet '{sheet_name}' to table '{target_table}'...")

                    total_rows = 0
                    for progress in import_excel_streaming(file_path, db_path, target_table, conn):
                        total_rows = progress
                        if progress % 50000 == 0:
                            print(f"  Progress: {progress} rows processed...")

                    print(f"Sheet '{sheet_name}' import completed. Loaded {total_rows} rows.")

                    sheet_file_name = f"{file_name}::{sheet_name}" if len(sheet_names) > 1 else file_name
                    record_import(conn, sheet_file_name, target_table, file_md5)
                    imported_tables.append(target_table)

        elif ext == '.numbers':
            try:
                from numbers_parser import Document
            except ImportError:
                raise ImportError("Please install 'numbers-parser' package to read .numbers files: pip install numbers-parser")

            print("Processing Mac Numbers file...")
            doc = Document(file_path)
            sheets = doc.sheets
            if not sheets:
                raise ValueError("No sheets found in the .numbers file.")

            for sheet in sheets:
                sheet_name = sheet.name
                if len(sheets) > 1:
                    safe_sheet_name = str(sheet_name).strip()
                    safe_sheet_name = re.sub(r'\W+', '_', safe_sheet_name).strip('_')
                    base_sheet_table = f"{table_name}_{safe_sheet_name}"
                else:
                    base_sheet_table = table_name

                target_table = get_unique_table_name(base_sheet_table)
                print(f"Processing Numbers sheet '{sheet_name}' to table '{target_table}'...")

                tables = sheet.tables
                if not tables:
                    print(f"No tables found in sheet '{sheet_name}', skipping.")
                    continue

                data = tables[0].rows(values_only=True)
                df = pd.DataFrame(data)

                df.dropna(how='all', inplace=True)
                df.dropna(axis=1, how='all', inplace=True)

                header_idx = find_header_row(df)
                if header_idx > 0:
                    new_header = df.iloc[header_idx]
                    df = df[header_idx+1:]
                    df.columns = new_header
                elif len(df) > 0:
                    df.columns = df.iloc[0]
                    df = df[1:]

                df.columns = clean_column_names(df.columns)
                _df_to_duckdb(conn, df, target_table, 'replace')
                print(f"Sheet '{sheet_name}' import completed. Loaded {len(df)} rows.")

                sheet_file_name = f"{file_name}::{sheet_name}" if len(sheets) > 1 else file_name
                record_import(conn, sheet_file_name, target_table, file_md5)
                imported_tables.append(target_table)

        else:
            raise ValueError(f"Unsupported file format: {ext}")

        return imported_tables


async def import_from_url(
    url: str,
    db_path: str,
    table_name: str,
    source_format: str,
    auth_config=None
) -> str:
    """Import data from URL into DuckDB."""
    from scripts.url_data_source import URLDataSource, URLDataSourceConfig, AuthConfig

    config = URLDataSourceConfig(
        url=url,
        format=source_format,
        table_name=table_name,
        auth=auth_config
    )

    source = URLDataSource(config)

    logger.info(
        "开始从 URL 导入数据",
        url=url,
        format=source_format,
        table=table_name
    )

    try:
        records = await source.fetch_and_parse()

        if not records:
            logger.warning("URL 返回空数据", url=url)
            raise ValueError("URL returned empty data")

        # Import to DuckDB
        repo = DatabaseRepository(db_path)
        with repo.connection() as conn:
            init_meta_table(conn)

            df = pd.DataFrame(records)
            df.columns = clean_column_names(df.columns.tolist())
            conn.register('_tmp_df', df)
            conn.execute(f"CREATE OR REPLACE TABLE {table_name} AS SELECT * FROM _tmp_df")
            conn.unregister('_tmp_df')

            # Record import metadata
            auth_type = None
            if auth_config is not None:
                auth_type = auth_config.type
            record_url_import(conn, url, table_name, source_format, auth_type)

        logger.info(
            "URL 数据导入完成",
            table=table_name,
            rows=len(records)
        )

        return table_name

    except Exception as e:
        logger.error(
            "URL 数据导入失败",
            url=url,
            error=str(e),
            error_type=type(e).__name__
        )
        raise


def import_from_url_sync(
    url: str,
    db_path: str = "workspace.duckdb",
    table_name: Optional[str] = None,
    source_format: str = "json",
    auth_config=None
) -> str:
    """Synchronous wrapper for URL import."""
    return asyncio.run(import_from_url(
        url=url,
        db_path=db_path,
        table_name=table_name,
        source_format=source_format,
        auth_config=auth_config
    ))


async def refresh_url_source(db_path: str, table_name: str) -> bool:
    """Refresh a URL data source.

    Re-fetches data from the original URL and replaces the table content.
    """
    from scripts.url_data_source import URLDataSource, URLDataSourceConfig

    repo = DatabaseRepository(db_path)

    with repo.connection() as conn:
        row = conn.execute('''
            SELECT source_url, source_format
            FROM _data_skill_meta
            WHERE table_name = ? AND source_url IS NOT NULL
        ''', (table_name,)).fetchone()

        if not row:
            raise ValueError(f"Table '{table_name}' is not a URL data source")

        source_url = row[0]
        source_format = row[1]

        logger.info(
            "开始刷新 URL 数据源",
            table=table_name,
            url=source_url
        )

        # Create config (no auth - user must re-authenticate)
        config = URLDataSourceConfig(
            url=source_url,
            format=source_format,
            table_name=table_name
        )

        source = URLDataSource(config)
        records = await source.fetch_and_parse()

        if not records:
            logger.warning("刷新返回空数据", table=table_name)
            raise ValueError("Refresh returned empty data")

        # Clear table and re-insert
        df = pd.DataFrame(records)
        df.columns = clean_column_names(df.columns.tolist())
        conn.register('_tmp_df', df)
        conn.execute(f"CREATE OR REPLACE TABLE {table_name} AS SELECT * FROM _tmp_df")
        conn.unregister('_tmp_df')

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute('''
            UPDATE _data_skill_meta
            SET last_refresh_time = ?, last_used_time = ?
            WHERE table_name = ?
        ''', (now, now, table_name))
        conn.commit()

        logger.info(
            "URL 数据源刷新完成",
            table=table_name,
            rows=len(records)
        )

        return True


def refresh_url_source_sync(db_path: str, table_name: str) -> bool:
    """Synchronous wrapper for refresh_url_source."""
    return asyncio.run(refresh_url_source(db_path, table_name))


def list_url_sources(db_path: str) -> List[dict]:
    """List all URL data sources."""
    repo = DatabaseRepository(db_path)

    with repo.connection() as conn:
        init_meta_table(conn)
        return get_url_sources(conn)


if __name__ == "__main__":  # pragma: no cover
    parser = argparse.ArgumentParser(
        description="Import data from files or URLs into DuckDB"
    )
    subparsers = parser.add_subparsers(dest="command", help="Import commands")

    # File subparser (existing functionality)
    file_parser = subparsers.add_parser("file", help="Import from file (CSV/Excel)")
    file_parser.add_argument("input_file", help="Path to the input Excel or CSV file")
    file_parser.add_argument("--db", default="workspace.duckdb", help="Path to DuckDB database file")
    file_parser.add_argument("--table", default=None, help="Target table name (auto-generated if not provided)")

    # URL subparser
    url_parser = subparsers.add_parser("url", help="Import from URL")
    url_parser.add_argument("url", help="HTTP/HTTPS URL to fetch data from")
    url_parser.add_argument("--format", required=True, choices=["json", "csv"], help="Data format")
    url_parser.add_argument("--table", required=True, help="Target table name")
    url_parser.add_argument("--db", default="workspace.duckdb", help="Path to DuckDB database file")
    url_parser.add_argument("--auth-type", choices=["basic", "bearer"], default=None, help="Authentication type")
    url_parser.add_argument("--auth-user", help="Username for Basic Auth")
    url_parser.add_argument("--auth-password", help="Password for Basic Auth")
    url_parser.add_argument("--auth-token", help="Token for Bearer Auth")

    # Refresh subparser
    refresh_parser = subparsers.add_parser("refresh", help="Refresh URL data source")
    refresh_parser.add_argument("table", help="Table name to refresh")
    refresh_parser.add_argument("--db", default="workspace.duckdb", help="Path to DuckDB database file")

    # List subparser
    list_parser = subparsers.add_parser("list", help="List URL data sources")
    list_parser.add_argument("--db", default="workspace.duckdb", help="Path to DuckDB database file")

    args = parser.parse_args()

    # Handle backward compatibility: if no subcommand, treat as file import
    if args.command is None:
        # Legacy mode: positional argument is input_file
        if hasattr(args, 'input_file') and args.input_file:
            try:
                final_tables = import_to_sqlite(args.input_file, args.db, getattr(args, 'table', None))
                if isinstance(final_tables, list):
                    print(f"SUCCESS: Data imported into tables: {', '.join(final_tables)}")
                else:
                    print(f"SUCCESS: Data imported into table '{final_tables}'")
            except Exception as e:
                print(f"ERROR: {e}")
        else:
            parser.print_help()
        exit(0)

    if args.command == "file":
        try:
            final_tables = import_to_sqlite(args.input_file, args.db, args.table)
            if isinstance(final_tables, list):
                print(f"SUCCESS: Data imported into tables: {', '.join(final_tables)}")
            else:
                print(f"SUCCESS: Data imported into table '{final_tables}'")
        except Exception as e:
            print(f"ERROR: {e}")

    elif args.command == "url":
        try:
            # Build auth config if provided
            auth_config = None
            if args.auth_type == "basic":
                if not args.auth_user or not args.auth_password:
                    print("ERROR: --auth-user and --auth-password required for Basic Auth")
                    exit(1)
                from scripts.url_data_source import BasicAuthConfig
                auth_config = BasicAuthConfig(
                    username=args.auth_user,
                    password=args.auth_password
                )
            elif args.auth_type == "bearer":
                if not args.auth_token:
                    print("ERROR: --auth-token required for Bearer Auth")
                    exit(1)
                from scripts.url_data_source import BearerAuthConfig
                auth_config = BearerAuthConfig(token=args.auth_token)

            # Import from URL
            table_name = import_from_url_sync(
                url=args.url,
                db_path=args.db,
                table_name=args.table,
                source_format=args.format,
                auth_config=auth_config
            )

            # Get row count
            conn = duckdb.connect(args.db)
            row_count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            conn.close()

            print(f"Imported {row_count} rows into table '{table_name}' from {args.url}")

        except Exception as e:
            print(f"ERROR: {e}")

    elif args.command == "refresh":
        try:
            result = refresh_url_source_sync(args.db, args.table)

            # Get row count
            conn = duckdb.connect(args.db)
            row_count = conn.execute(f"SELECT COUNT(*) FROM {args.table}").fetchone()[0]
            conn.close()

            print(f"Refreshed table '{args.table}' with {row_count} rows")

        except Exception as e:
            print(f"ERROR: {e}")

    elif args.command == "list":
        try:
            sources = list_url_sources(args.db)

            if not sources:
                print("No URL data sources found.")
            else:
                # Print formatted table
                print(f"{'Table':<20} {'URL':<40} {'Format':<8} {'Auth':<8}")
                print("-" * 80)
                for source in sources:
                    print(f"{source['table_name']:<20} {source['source_url']:<40} {source['source_format']:<8} {(source['auth_type'] or 'none'):<8}")

        except Exception as e:
            print(f"ERROR: {e}")
